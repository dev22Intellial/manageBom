from django import forms
from .models import BOMFile
import os
import openpyxl

class BOMUploadForm(forms.ModelForm):
    class Meta:
        model = BOMFile
        fields = ['name', 'file']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'file': forms.FileInput(attrs={'class': 'form-control', 'accept': '.xlsx'}),
        }

    def clean_file(self):
        file = self.cleaned_data.get('file', False)
        if not file:
            return file

        # 1. Check file extension
        ext = os.path.splitext(file.name)[1]
        if not ext.lower() == '.xlsx':
            raise forms.ValidationError('Only .xlsx files are allowed.')

        required_columns = [
            'Reference designators', 
            'Quantity', 
            'Identified MPN', 
            'Identified manufacturer'
        ]

        try:
            # Use a temporary in-memory file to avoid saving before validation
            file.seek(0)
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Check if sheet is empty
            if sheet.max_row < 1:
                raise forms.ValidationError("The uploaded file is empty.")

            headers = [cell.value for cell in sheet[1]]
            
            missing_columns = [col for col in required_columns if col not in headers]

            if missing_columns:
                # User-specified error message for missing columns
                raise forms.ValidationError(
                    f"The file does not have one of the required columns ({', '.join(required_columns)})."
                )
            
            # 3. Check for at least one data row
            has_data_row = False
            for row_index in range(2, sheet.max_row + 1):
                row_data = [cell.value for cell in sheet[row_index]]
                if any(cell is not None for cell in row_data): # Check if row is not entirely empty
                    has_data_row = True
                    break
            
            if not has_data_row:
                raise forms.ValidationError("The uploaded file contains no valid BOM entries after the header row.")

        except forms.ValidationError as e:
            # Re-raise specific validation errors
            raise e
        except Exception:
            # Catch all other exceptions for corrupted or unreadable files
            raise forms.ValidationError("Could not read the file. It may be corrupted or not a valid XLSX file.")
        
        # Rewind file for storage
        file.seek(0)
        return file